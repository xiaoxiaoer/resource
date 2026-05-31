"""
Excel 公式校验模块

对新模板（Format C）中的公式单元格进行自动校验：
1. 分别以 data_only=True/False 加载 workbook
2. 提取输入值，按依赖链顺序重算公式
3. 对比计算结果与 Excel 缓存值
"""

import math
from pathlib import Path

import openpyxl


def _num(val):
    if val is None or val == '':
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0


def _str(val):
    if val is None:
        return ''
    return str(val).strip()


def _ceiling(value, significance=1):
    if significance == 0:
        return 0
    return math.ceil(value / significance) * significance


def _values_close(a, b):
    if a is None or b is None:
        return a is None and b is None
    if isinstance(a, str) or isinstance(b, str):
        return str(a).strip() == str(b).strip()
    if a == 0 and b == 0:
        return True
    if abs(a - b) <= 0.5:
        return True
    return abs(a - b) / max(abs(a), abs(b)) < 0.001


def _is_excel_error(val):
    return isinstance(val, str) and val.startswith('#')


def _make_detail(cell, sheet, label, formula, computed, cached):
    if _is_excel_error(cached):
        match = None
    elif cached is None or (isinstance(cached, str) and cached.strip() == '' and not isinstance(computed, str)):
        match = None
    else:
        match = _values_close(computed, cached)
    return {
        'cell': cell,
        'sheet': sheet,
        'label': label,
        'formula': formula,
        'computed': round(computed, 10) if isinstance(computed, float) else computed,
        'cached': round(cached, 10) if isinstance(cached, float) else cached,
        'match': match,
    }


def _verify_evaluation_sheet(ws_v, ws_f):
    details = []
    read = lambda r, c: ws_v.cell(row=r, column=c).value
    fstr = lambda r, c: ws_f.cell(row=r, column=c).value
    label = lambda r, c=2: _str(read(r, c)).split('\n')[0]
    is_formula = lambda r, c: (v := fstr(r, c)) and isinstance(v, str) and v.startswith('=')

    # ── 输入值 ──
    C7  = _num(read(7, 3));  C9  = _num(read(9, 3))
    C10 = _num(read(10, 3)); C11 = _num(read(11, 3))
    C12 = _num(read(12, 3)); C13 = _num(read(13, 3))
    C14 = _num(read(14, 3))
    C15_raw = read(15, 3);   C15 = _num(C15_raw)
    C17 = _num(read(17, 3)); C18 = _num(read(18, 3))

    B12 = _str(read(12, 2)); F11 = _str(read(11, 6))

    G7  = _num(read(7, 7));  G9  = _num(read(9, 7))
    G10 = _num(read(10, 7)); G11 = _num(read(11, 7))
    G12 = _num(read(12, 7)); G13 = _num(read(13, 7))
    G14 = _num(read(14, 7)); G15 = _str(read(15, 7))
    G16 = _num(read(16, 7))

    D33 = _str(read(33, 4)); D34 = _num(read(34, 4))
    D35 = _str(read(35, 4)); D36 = _str(read(36, 4))

    def add(r, c, computed):
        f = fstr(r, c)
        if not (f and isinstance(f, str) and f.startswith('=')):
            return
        details.append(_make_detail(
            ws_v.cell(row=r, column=c).coordinate,
            '项目评估计算', label(r), f, computed, read(r, c),
        ))

    # ── Group 1: 纯输入 ──

    C20 = C9 + C10 + G9
    add(20, 3, C20)

    first22 = 0 if C13 == 0 else (C10 + C9) / C13 * 0.03
    C22 = first22 + G12 * G14 * (G13 * 0.7) * 0.05
    add(22, 3, C22)

    if C15 == 0 or C15_raw is None:
        C23 = min(5, _ceiling(G14 / 12, 1))
    else:
        C23 = min(5, _ceiling(C15 / 12, 1))
    add(23, 3, C23)

    first27 = 0 if C13 == 0 else (C9 + C10) / C13
    rev27 = G12 * G13 * 0.7 * G14
    share27 = (rev27 - G9) * (G16 * 0.01) if G15 == '回本后分润' else rev27 * (G16 * 0.01)
    C27 = first27 + rev27 - share27
    add(27, 3, C27)

    # ── Group 2 ──

    def _tax(inv_type, vat, base):
        if inv_type == '增值税专用发票':
            if vat == 3:  return base * 0.06
            if vat == 5:  return base * 0.04
            return 0
        return base * 0.09

    C21 = _tax(B12, C12, C20) + _tax(F11, G11, C20)
    add(21, 3, C21)

    disc_map = {1: 0.91, 2: 0.87, 3: 0.82, 4: 0.77, 5: 0.72}
    E42 = disc_map.get(int(C23), '请手动输入') if isinstance(C23, (int, float)) else '请手动输入'
    add(42, 5, E42)

    if isinstance(E42, (int, float)):
        E43 = C27 * E42
    else:
        E43 = 0
    add(43, 5, E43)

    # ── Group 3 ──

    denom45 = G7 + C7
    if denom45 == 0:
        E45 = 0
    else:
        E45 = max(0, min(1, round((E43 - (G10 + C11 + C21 + C22)) / denom45, 2)))
    add(45, 5, E45)

    C8 = C7 * E45
    add(8, 3, C8)
    G8 = G7 * E45
    add(8, 7, G8)

    # ── Group 4 ──

    C24 = C21 + C22 + C11 + C10 + C7 + G7 + G10
    add(24, 3, C24)
    C25_cost = C21 + C22 + C11 + C10 + C8 + G8 + G10
    add(25, 3, C25_cost)
    C19 = C24 / C27 if C27 != 0 else 0
    add(19, 3, C19)

    try:
        if C15_raw is None or _str(C15_raw) == '':
            C28 = C27 / G14 if G14 != 0 else 0
        else:
            C28 = C27 / C15 if C15 != 0 else 0
    except ZeroDivisionError:
        C28 = 0
    add(28, 3, C28)

    C26 = C28 / (C17 + C18) if (C17 + C18) != 0 else 0
    add(26, 3, C26)

    # ── Group 5: 风险评分 ──

    D32 = C26
    add(32, 4, D32)
    E32 = 0 if D32 <= 0.4 else (3 if D32 <= 0.6 else (5 if D32 <= 0.7 else 20))
    add(32, 5, E32)

    e33_map = {'≥1年': 0, '0.5年-1年': 5, '0年-0.5年': 10, '<0年': 30}
    E33 = e33_map.get(D33, '')
    add(33, 5, E33)

    if D34 == '' or D34 is None:
        E34 = ''
    elif D34 >= 12: E34 = 0
    elif D34 >= 6:  E34 = 3
    elif D34 >= 3:  E34 = 10
    else:           E34 = 20
    add(34, 5, E34)

    e35_map = {'中大型': 0, '小型': 1, '微型': 10, '个体': 10}
    E35 = e35_map.get(D35, '')
    add(35, 5, E35)

    e36_map = {'ETC+自有小程序支付': 10, 'ETC支付': 7, '自有小程序支付': 3, '无ETC、自有小程序支付': 0}
    E36 = e36_map.get(D36, '')
    add(36, 5, E36)

    D37 = C10 + C8 + G8
    add(37, 4, D37)
    E37 = 1 if D37 <= 100000 else (3 if D37 <= 300000 else (5 if D37 <= 500000 else 10))
    add(37, 5, E37)

    E38 = 100 - sum(x for x in [E32, E33, E34, E35, E36, E37] if isinstance(x, (int, float)))
    add(38, 5, E38)

    if   E38 < 60: E39 = '总经理权限'
    elif E38 < 70: E39 = '副总经理权限'
    elif E38 < 80: E39 = '分管总监权限'
    else:          E39 = '大区经理权限'
    add(39, 5, E39)

    # ── Group 6: 综合结果 ──

    E44 = (E43 - C24) / E43 if E43 != 0 else 0
    add(44, 5, E44)

    if   E45 >= 1:    E46 = '无需审批'
    elif E45 >= 0.85: E46 = '大区经理'
    elif E45 >= 0.75: E46 = '分管总监'
    else:             E46 = '副总经理'
    add(46, 5, E46)

    E47 = (E43 - C25_cost) / E43 if E43 != 0 else 0
    add(47, 5, E47)

    if   E47 < 0:    F47 = '总经理权限'
    elif E47 < 0.05: F47 = '副总经理权限'
    elif E47 < 0.1:  F47 = '分管总监权限'
    else:            F47 = '大区经理权限'
    add(47, 6, F47)

    if E39 == F47:
        E48 = E39
    elif '总经理权限' in (E39, F47):   E48 = '总经理权限'
    elif '副总经理权限' in (E39, F47): E48 = '副总经理权限'
    elif '分管总监权限' in (E39, F47): E48 = '分管总监权限'
    else:                              E48 = '大区经理权限'
    add(48, 5, E48)

    E49 = C8 + G8
    add(49, 5, E49)

    if E47 < 0:
        E50 = '0'
    else:
        E50 = E43 * E47 * 0.15
    add(50, 5, E50)

    return details


def _verify_cooperation_sheet(ws_v, ws_f):
    details = []

    for col, col_label in [(3, '月均临停收入'), (4, '月均月租收入')]:
        formula = ws_f.cell(row=25, column=col).value
        if not (formula and isinstance(formula, str) and formula.startswith('=')):
            continue

        vals = [_num(ws_v.cell(row=r, column=col).value) for r in range(13, 25)]
        pos = [v for v in vals if v > 0]
        computed = sum(pos) / len(pos) if pos else 0

        details.append(_make_detail(
            ws_v.cell(row=25, column=col).coordinate,
            '合作车场信息收集表', col_label, formula, computed,
            ws_v.cell(row=25, column=col).value,
        ))

    return details


def verify_formulas(file_path: str) -> dict:
    path = Path(file_path)
    if not path.exists():
        return {'has_formulas': False, 'total': 0, 'matched': 0, 'mismatched': 0, 'no_cache': 0, 'details': []}

    wb_v = openpyxl.load_workbook(str(path), data_only=True)
    wb_f = openpyxl.load_workbook(str(path), data_only=False)

    if '项目评估计算' not in wb_v.sheetnames:
        return {'has_formulas': False, 'total': 0, 'matched': 0, 'mismatched': 0, 'no_cache': 0, 'details': []}

    details = _verify_evaluation_sheet(wb_v['项目评估计算'], wb_f['项目评估计算'])

    if '合作车场信息收集表' in wb_v.sheetnames:
        details.extend(_verify_cooperation_sheet(wb_v['合作车场信息收集表'], wb_f['合作车场信息收集表']))

    matched = sum(1 for d in details if d['match'] is True)
    mismatched = sum(1 for d in details if d['match'] is False)
    no_cache = sum(1 for d in details if d['match'] is None)

    return {
        'has_formulas': len(details) > 0,
        'total': len(details),
        'matched': matched,
        'mismatched': mismatched,
        'no_cache': no_cache,
        'details': details,
    }
