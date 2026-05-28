"""
Excel 解析器 — 解析资源置换运营测算表

支持两种 sheet 结构：
1. 金额券测算（停车券业务）：含"项目基本信息" + "测算工具（金额券）"
2. 时长采买测算（车位置换业务）：含"折扣采买测算表"
"""

import json
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

# openpyxl 已知 bug：新版 Excel 的 DataValidation 带 `id` 属性，openpyxl 不识别
# monkey-patch 让它忽略该属性
try:
    _orig_init = openpyxl.worksheet.datavalidation.DataValidation.__init__
    def _patched_init(self, *args, **kwargs):
        kwargs.pop('id', None)
        _orig_init(self, *args, **kwargs)
    openpyxl.worksheet.datavalidation.DataValidation.__init__ = _patched_init
except Exception:
    pass


def _serial_to_date(serial: float | int | None) -> str | None:
    if serial is None:
        return None
    try:
        # Excel serial date: 1 = 1900-01-01, but Excel has a leap year bug (1900-02-29)
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=int(serial))).strftime('%Y-%m-%d')
    except (ValueError, OverflowError):
        return str(serial)


def _safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _safe_int(val) -> int | None:
    f = _safe_float(val)
    return int(f) if f is not None else None


def _normalize_vat(val) -> float | None:
    """标准化增值税率为小数（如 0.09）。输入可能是整数 9 或小数 0.09。"""
    val = _safe_float(val)
    if val is None:
        return None
    return val / 100.0 if val > 1 else val


def _cell_str(ws, row: int, col: int) -> str | None:
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    return str(val).strip() or None


def _cell_num(ws, row: int, col: int) -> float | None:
    return _safe_float(ws.cell(row=row, column=col).value)


def parse_survey_form(ws) -> dict:
    """解析「新车场调研表」sheet — 横向表头+值布局，提取补充项目信息"""
    result = {}

    # Row 4 表头 → Row 5 值, Row 6 表头 → Row 7 值
    for header_row, value_row in [(4, 5), (6, 7)]:
        header_map = {}
        for col in range(1, ws.max_column + 1):
            label = _cell_str(ws, header_row, col)
            if label:
                header_map[label] = col
        for target_key, keywords in [
            ('city', ['所在城市']),
            ('district', ['行政区']),
            ('car_park_name', ['项目名称']),
            ('business_nature', ['业态']),
            ('cooperation_model', ['合作模式']),
            ('parking_spaces', ['总车位数']),
            ('parking_lot_type', ['车位类型']),
            ('parking_fee_rule', ['临停收费规则']),
            ('monthly_card_price', ['线下月卡标准', '月卡标准']),
            ('main_income_source', ['收入来源']),
            ('busy_period', ['忙时时段']),
            ('occupancy_rate', ['占用率']),
            ('existing_monthly_cards', ['现有月卡数']),
            ('has_charging_pile', ['充电桩']),
        ]:
            if target_key in result:
                continue
            for kw in keywords:
                for label, col in header_map.items():
                    if kw in label:
                        val = ws.cell(row=value_row, column=col).value
                        if target_key == 'parking_spaces':
                            result[target_key] = _safe_int(_safe_float(val))
                        elif target_key in ('monthly_card_price', 'occupancy_rate'):
                            result[target_key] = _safe_float(val)
                        elif val is not None:
                            result[target_key] = str(val).strip() or None
                        break
                else:
                    continue
                break

    # 合作条款区（Row 12 附近）
    for row in range(12, 15):
        for col in range(1, 8):
            label = _cell_str(ws, row, col)
            if label and '允许粘贴' in label:
                result['allow_posting'] = _cell_str(ws, row + 1, col)
                break

    return result


def parse_cooperation_info(ws) -> dict:
    """解析「合作车场信息收集表」sheet — 标签在 col3，值在 col4（与旧格式偏移+1）"""
    result = {}

    # 车场名称、车场地址：标签在 col 2，值在 col 3
    name_row = _find_label_row(ws, 2, '车场名称', 1, 5)
    result['car_park_name'] = _cell_str(ws, name_row, 3) if name_row else None
    addr_row = _find_label_row(ws, 2, '车场地址', 1, 5)
    result['car_park_address'] = _cell_str(ws, addr_row, 3) if addr_row else None

    # 其余字段：标签在 col 3，值在 col 4
    field_defs = [
        ('property_type',        ['合作客户', '产权方']),
        ('contract_expire_date', ['承包到期']),
        ('parking_fee_rule',     ['收费规则']),
        ('allow_posting',        ['张贴物料', '允许张贴']),
        ('parking_spaces',       ['车位数量', '车位']),
    ]
    for key, keywords in field_defs:
        for kw in keywords:
            row = _find_label_row(ws, 3, kw, 4, 25)
            if row:
                if key == 'contract_expire_date':
                    result[key] = _serial_to_date(_cell_num(ws, row, 4)) or _cell_str(ws, row, 4)
                elif key == 'parking_spaces':
                    result[key] = _safe_int(_cell_num(ws, row, 4))
                else:
                    result[key] = _cell_str(ws, row, 4)
                break

    # 模板占位符清理
    for key in ('contract_expire_date',):
        val = result.get(key)
        if val and not any(c.isdigit() for c in str(val)):
            result[key] = None

    # 月度收入：col 2 月份，col 3 临停，col 4 月租
    header_row = _find_label_row(ws, 2, '月份', 10, 25)
    if header_row:
        monthly_income = []
        for row in range(header_row + 1, header_row + 13):
            month = _safe_int(_cell_num(ws, row, 2))
            if month is None:
                continue
            temp = _cell_num(ws, row, 3)
            ticket = _cell_num(ws, row, 4)
            if temp is None and ticket is None:
                continue
            monthly_income.append({
                'month': month,
                'temp_income': temp,
                'ticket_income': ticket,
            })
        result['monthly_income'] = monthly_income

        avg_row = _find_label_row(ws, 2, '月均', header_row, header_row + 15)
        if avg_row:
            result['monthly_avg_temp'] = _cell_num(ws, avg_row, 3)
            result['monthly_avg_ticket'] = _cell_num(ws, avg_row, 4)

    return result


def parse_evaluation_calc(ws) -> tuple[dict, dict | None]:
    """解析「项目评估计算」sheet — 左侧停车券(col2-4) + 右侧车位(col6-8) + 风险评估"""
    # 构建 col 2 标签→行号 映射（左侧停车券区）
    label_rows_left = {}
    for row in range(1, min(ws.max_row + 1, 30)):
        label = _cell_str(ws, row, 2)
        if label:
            label_rows_left[label] = row

    def _find_left(keyword):
        for label, row in label_rows_left.items():
            if keyword in label:
                return row
        return None

    # 左侧停车券 → calculation_tool
    voucher = {
        'department': None,
        'submitter': None,
        'opportunity_name': None,
    }

    calc_fields = [
        ('equipment_service_amount', '设备/置换金额'),
        ('cash_purchase_amount',     '现金采买'),
        ('business_fee',             '业务费'),
        ('vat_rate',                 '增值税'),
        ('discount_rate',            '签约折扣率'),
        ('contract_months',          '合同年限', True),
        ('voucher_consume_months',   '消耗年限', True),
        ('monthly_temp_income',      '月均临停'),
        ('monthly_ticket_income',    '月均月租'),
        ('contract_amount',          '合同金额'),
        ('voucher_total_value',      '停车券总价值'),
        ('voucher_monthly_share',    '每月消耗金额'),
        ('monthly_consume_ratio',    '月均消耗比例'),
        ('actual_purchase_discount', '实际折扣率'),
    ]
    for item in calc_fields:
        key, kw = item[0], item[1]
        as_int = item[2] if len(item) > 2 else False
        row = _find_left(kw)
        if row is None:
            voucher[key] = None
            continue
        v = _cell_num(ws, row, 3)
        if as_int:
            v = _safe_int(v)
        if key == 'vat_rate':
            v = _normalize_vat(v)
        voucher[key] = v

    # 风险评估区（rows 30-34, cols 2-5）
    evaluation_scores = []
    for row in range(30, 35):
        dimension = _cell_str(ws, row, 2)
        if dimension:
            evaluation_scores.append({
                'category': dimension,
                'value': str(ws.cell(row=row, column=4).value if ws.cell(row=row, column=4).value is not None else ''),
                'score': str(ws.cell(row=row, column=5).value if ws.cell(row=row, column=5).value is not None else ''),
            })
    voucher['evaluation_scores'] = evaluation_scores
    voucher['risk_rating'] = _cell_str(ws, 35, 5)

    # 综合结果区（col 5, rows 38-44）
    voucher['overall_assessment'] = _cell_str(ws, 42, 5)

    # 右侧车位 → spot_exchange_calc
    label_rows_right = {}
    for row in range(5, min(ws.max_row + 1, 20)):
        label = _cell_str(ws, row, 6)
        if label:
            label_rows_right[label] = row

    def _find_right(keyword):
        for label, row in label_rows_right.items():
            if keyword in label:
                return row
        return None

    spot_fields = [
        ('equipment_amount',     ['设备/置换金额']),
        ('vat_rate',             ['增值税']),
        ('monthly_card_fee',     ['对外月租价格']),
        ('replacement_spaces',   ['置换车位数量']),
        ('profit_share_ratio',   ['回本后甲方分润比例']),
        ('contract_months',      ['合同年限']),
    ]
    spot = {}
    for key, keywords in spot_fields:
        row = None
        for kw in keywords:
            row = _find_right(kw)
            if row:
                break
        if row is None:
            spot[key] = None
            continue
        v = _cell_num(ws, row, 8)
        if key == 'vat_rate':
            v = _normalize_vat(v)
        elif key == 'replacement_spaces':
            v = _safe_int(v)
        elif key == 'contract_months':
            v = _safe_int(v)
        spot[key] = v

    # 仅当右侧有实际数据时返回 spot 字典
    spot_has_data = any(v is not None for k, v in spot.items()
                        if k in ('equipment_amount', 'replacement_spaces',
                                 'monthly_card_fee', 'profit_share_ratio'))
    return voucher, spot if spot_has_data else None


def _find_label_row(ws, col: int, keyword: str, start: int = 1, end: int = 30) -> int | None:
    """在指定列中查找包含关键词的第一个行号"""
    for row in range(start, end + 1):
        val = ws.cell(row=row, column=col).value
        if val and keyword in str(val):
            return row
    return None


def parse_project_info(ws) -> dict:
    """解析 Sheet 1: 项目基本信息（通过标签关键词查找，兼容不同版本格式）"""
    result = {
        'car_park_name': _cell_str(ws, 2, 2),
        'car_park_address': _cell_str(ws, 3, 2),
    }

    # 通过 col 2 标签关键词查找各行（兼容标签文字变化）
    field_defs = [
        ('property_type',     ['合作客户', '产权方']),
        ('contract_expire_date', ['承包到期']),
        ('parking_fee_rule',  ['收费规则']),
        ('allow_posting',     ['张贴物料', '允许张贴']),
        ('has_own_channel',   ['自有', 'ETC']),
        ('parking_spaces',    ['车位数量', '车位']),
    ]
    for key, keywords in field_defs:
        for kw in keywords:
            row = _find_label_row(ws, 2, kw, 4, 15)
            if row:
                if key == 'contract_expire_date':
                    result[key] = _serial_to_date(_cell_num(ws, row, 3)) or _cell_str(ws, row, 3)
                elif key == 'parking_spaces':
                    result[key] = _safe_int(_cell_num(ws, row, 3))
                else:
                    result[key] = _cell_str(ws, row, 3)
                break

    # 模板占位符清理
    for key in ('contract_expire_date',):
        val = result.get(key)
        if val and not any(c.isdigit() for c in str(val)):
            result[key] = None

    # 月度收入 — 查找「月份」表头行，只保留有实际数据的月份
    header_row = _find_label_row(ws, 1, '月份', 10, 20)
    if header_row:
        monthly_income = []
        for row in range(header_row + 1, header_row + 13):
            month = _safe_int(_cell_num(ws, row, 1))
            if month is None:
                continue
            temp = _cell_num(ws, row, 2)
            ticket = _cell_num(ws, row, 3)
            if temp is None and ticket is None:
                continue
            monthly_income.append({
                'month': month,
                'temp_income': temp,
                'ticket_income': ticket,
            })
        result['monthly_income'] = monthly_income

        avg_row = _find_label_row(ws, 1, '月均', header_row, header_row + 15)
        if avg_row:
            result['monthly_avg_temp'] = _cell_num(ws, avg_row, 2)
            result['monthly_avg_ticket'] = _cell_num(ws, avg_row, 3)

    return result


def parse_project_info_collection(ws) -> dict:
    """解析「项目信息收集表」sheet（车位置换业务的项目信息）— 通过标签关键词查找，兼容行偏移"""
    result = {}

    # 车场名称、车场地址：标签在 col 1，值在 col 2
    name_row = _find_label_row(ws, 1, '车场名称', 1, 5)
    result['car_park_name'] = _cell_str(ws, name_row, 2) if name_row else _cell_str(ws, 2, 2)
    addr_row = _find_label_row(ws, 1, '车场地址', 1, 5)
    result['car_park_address'] = _cell_str(ws, addr_row, 2) if addr_row else _cell_str(ws, 3, 2)

    # 其余字段：标签在 col 2，值在 col 3，通过关键词查找
    field_defs = [
        ('business_nature',      ['车场业态', '业态']),
        ('property_type',        ['合作客户', '产权方']),
        ('contract_expire_date', ['承包到期']),
        ('parking_fee_rule',     ['收费规则']),
        ('occupancy_rate',       ['车位占用率', '占用率']),
        ('allow_posting',        ['张贴物料', '允许张贴']),
        ('settlement_mode',      ['结算模式', '结算']),
    ]
    for key, keywords in field_defs:
        for kw in keywords:
            row = _find_label_row(ws, 2, kw, 4, 25)
            if row:
                if key == 'contract_expire_date':
                    result[key] = _serial_to_date(_cell_num(ws, row, 3)) or _cell_str(ws, row, 3)
                else:
                    result[key] = _cell_str(ws, row, 3)
                break

    # 模板占位符清理
    for key in ('contract_expire_date',):
        val = result.get(key)
        if val and not any(c.isdigit() for c in str(val)):
            result[key] = None

    # 月度收入 — 查找「月份」表头行，只保留有实际数据的月份
    header_row = _find_label_row(ws, 1, '月份', 10, 25)
    if not header_row:
        header_row = 14
    monthly_income = []
    for row in range(header_row + 1, header_row + 13):
        month = _safe_int(_cell_num(ws, row, 1))
        if month is None:
            continue
        temp = _cell_num(ws, row, 2)
        ticket = _cell_num(ws, row, 3)
        if temp is None and ticket is None:
            continue
        monthly_income.append({
            'month': month,
            'temp_income': temp,
            'ticket_income': ticket,
        })
    result['monthly_income'] = monthly_income

    return result


def parse_spot_exchange_calc(ws) -> dict:
    """解析「车位置换测算表」sheet（车位置换业务测算数据）"""
    result = {
        'department': _cell_str(ws, 3, 3),
        'submitter': _cell_str(ws, 3, 5),
        'opportunity_name': _cell_str(ws, 4, 3),
        'crm_quote_no': _cell_str(ws, 5, 3),
        'crm_m_factor': _cell_num(ws, 5, 5),
        'crm_settlement_total': _cell_num(ws, 6, 3),
        'crm_service_fee': _cell_num(ws, 6, 5),
        'equipment_amount': _cell_num(ws, 9, 4),
        'vat_rate': _normalize_vat(_cell_num(ws, 10, 4)),
        'monthly_card_fee': _cell_num(ws, 11, 4),
        'replacement_spaces': _safe_int(_cell_num(ws, 12, 4)),
        'profit_share_ratio': _cell_num(ws, 13, 4),
        'contract_months': _safe_int(_cell_num(ws, 14, 4)),
        'per_space_monthly_income': _cell_num(ws, 15, 4),
        'purchase_unit_price': _cell_num(ws, 16, 4),
        'tax_cost': _cell_num(ws, 17, 4),
        'total_cost': _cell_num(ws, 18, 4),
        'total_revenue': _cell_num(ws, 19, 4),
        'our_profit': _cell_num(ws, 20, 4),
        'customer_profit': _cell_num(ws, 21, 4),
    }
    return result


def parse_calculation_tool(ws) -> dict:
    """解析 Sheet 2: 测算工具（金额券）— 通过标签关键词查找，兼容行偏移"""
    # 构建 col 2 标签→行号 映射
    label_rows = {}
    for row in range(1, min(ws.max_row + 1, 30)):
        label = _cell_str(ws, row, 2)
        if label:
            label_rows[label] = row

    def _find(keyword):
        for label, row in label_rows.items():
            if keyword in label:
                return row
        return None

    def _num(keyword, col, as_int=False):
        row = _find(keyword)
        if row is None:
            return None
        v = _cell_num(ws, row, col)
        return _safe_int(v) if as_int else v

    def _str(keyword, col):
        row = _find(keyword)
        return _cell_str(ws, row, col) if row else None

    # 业务员信息
    result = {
        'department': _str('提交部门', 3),
        'submitter': _str('提交部门', 5),
        'opportunity_name': _str('商机名称', 3),
        'installment_deduction': _str('商机名称', 5),
        'crm_service_fee': _cell_num(ws, 7, 5),
    }

    # 计算字段（col 2 标签, col 4 数值）
    calc_fields = [
        ('equipment_service_amount', '设备、服务置换'),
        ('cash_purchase_amount',     '现金采买'),
        ('business_fee',             '业务费'),
        ('discount_rate',            '采买折扣'),
        ('contract_months',          '合同有效年限', True),
        ('voucher_consume_months',   '券消耗年限', True),
        ('vat_rate',                 '增值税'),
        ('monthly_temp_income',      '月均临停'),
        ('monthly_ticket_income',    '月均月租'),
        ('monthly_total_income',     '月均收入'),
        ('contract_amount',          '合同金额'),
        ('voucher_total_value',      '停车券总价值'),
        ('voucher_monthly_share',    '停车券月均分摊'),
        ('project_gross_profit',     '项目整体毛利'),
    ]
    for item in calc_fields:
        key, kw = item[0], item[1]
        as_int = item[2] if len(item) > 2 else False
        v = _num(kw, 4, as_int=as_int)
        if key == 'vat_rate':
            v = _normalize_vat(v)
        result[key] = v

    # 付款方式（字符串字段）
    result['payment_method'] = _str('付款方式', 4)

    # 评估区（右侧，rows 4-6 相对固定）
    result['monthly_consume_ratio'] = _cell_num(ws, 4, 8)
    result['consume_ratio_status'] = _cell_str(ws, 4, 9)
    result['actual_purchase_discount'] = _cell_num(ws, 5, 8)
    result['discount_status'] = _cell_str(ws, 5, 9)
    result['discount_range'] = _cell_str(ws, 5, 10)
    result['overall_assessment'] = _cell_str(ws, 6, 8)

    # 评分表（rows 22-26, cols G-L）— 通过标签查找起始行
    score_header_row = _find_label_row(ws, 7, '所属类别', 18, 30)
    if score_header_row:
        evaluation_scores = []
        for r in range(score_header_row + 1, score_header_row + 5):
            left_cat = _cell_str(ws, r, 7)
            left_val = _cell_str(ws, r, 8)
            left_score = _cell_str(ws, r, 9)
            if left_cat:
                evaluation_scores.append({
                    'category': left_cat,
                    'value': left_val,
                    'score': left_score,
                })
            right_cat = _cell_str(ws, r, 10)
            right_val = _cell_str(ws, r, 11)
            right_score = _cell_str(ws, r, 12)
            if right_cat and '风险评分' in right_cat:
                result['risk_rating'] = right_score
            elif right_cat:
                evaluation_scores.append({
                    'category': right_cat,
                    'value': right_val,
                    'score': right_score,
                })
        result['evaluation_scores'] = evaluation_scores

    return result


def parse_discount_purchase(ws) -> dict | None:
    """解析 Sheet 3: 折扣采买测算表（车位置换业务）"""
    project_name = _cell_str(ws, 2, 2)
    if not project_name or 'XXXXXXX' in (project_name or ''):
        return None

    result = {
        'project_name': project_name,
        'parameters': {},
    }

    # 右侧参数表（rows 4-21, columns N-R）
    for row in range(4, 22):
        seq = _cell_str(ws, row, 14)   # N: 序号
        name = _cell_str(ws, row, 15)  # O: 项目
        value = ws.cell(row=row, column=16).value  # P: 数值
        unit = _cell_str(ws, row, 17)  # Q: 单位
        note = _cell_str(ws, row, 18)  # R: 备注
        if name:
            result['parameters'][name] = {
                'value': value,
                'unit': unit,
                'note': note,
            }

    # 左侧测算数据
    result['total_original_amount'] = _cell_num(ws, 4, 4)
    result['total_aike_amount'] = _cell_num(ws, 5, 4)

    return result


def parse_audit_excel(file_path: str | Path) -> dict:
    """
    解析资源置换运营测算表。

    Returns:
        {
            'file_name': str,
            'project_info': dict,        # Sheet 1
            'calculation_tool': dict,    # Sheet 2（停车券业务）
            'discount_purchase': dict,   # Sheet 3（车位置换业务）
            'business_type': str,        # 自动判断: parking_voucher / spot_exchange / both
        }
    """
    path = Path(file_path)
    wb = openpyxl.load_workbook(str(path), data_only=True)

    result = {'file_name': path.name}

    # === 格式C 检测：新模板（新车场调研表 + 合作车场信息收集表 + 项目评估计算）===
    is_format_c = '新车场调研表' in wb.sheetnames and '合作车场信息收集表' in wb.sheetnames

    if is_format_c:
        # 项目信息：合并两个 sheet
        project_info = {}
        if '新车场调研表' in wb.sheetnames:
            project_info.update(parse_survey_form(wb['新车场调研表']))
        if '合作车场信息收集表' in wb.sheetnames:
            coop = parse_cooperation_info(wb['合作车场信息收集表'])
            project_info.update(coop)  # 合作信息覆盖 survey 重复字段（更权威）
        result['project_info'] = project_info

        # 测算数据：单 sheet 左右分区
        if '项目评估计算' in wb.sheetnames:
            voucher_calc, spot_calc = parse_evaluation_calc(wb['项目评估计算'])
            if voucher_calc:
                result['calculation_tool'] = voucher_calc
            if spot_calc:
                result['spot_exchange_calc'] = spot_calc
    else:
        # === 格式A/B：原有模板 ===
        # 项目信息
        if '1.项目基本信息' in wb.sheetnames:
            result['project_info'] = parse_project_info(wb['1.项目基本信息'])
        elif '项目信息收集表' in wb.sheetnames:
            result['project_info'] = parse_project_info_collection(wb['项目信息收集表'])

        # 停车券业务: 测算工具（金额券）
        calc_sheet_name = None
        for name in wb.sheetnames:
            if '测算工具' in name and '金额券' in name:
                calc_sheet_name = name
                break
        if calc_sheet_name:
            result['calculation_tool'] = parse_calculation_tool(wb[calc_sheet_name])

        # 车位置换业务: 折扣采买测算表
        discount_sheet = None
        for name in wb.sheetnames:
            if '折扣采买' in name or '时长采买' in name:
                discount_sheet = name
                break
        if discount_sheet:
            parsed = parse_discount_purchase(wb[discount_sheet])
            if parsed:
                result['discount_purchase'] = parsed

        # 车位置换业务: 车位置换测算表（非「运营」版本）
        spot_calc_sheet = None
        for name in wb.sheetnames:
            if '车位置换测算表' in name and '运营' not in name:
                spot_calc_sheet = name
                break
        if spot_calc_sheet:
            result['spot_exchange_calc'] = parse_spot_exchange_calc(wb[spot_calc_sheet])

    # 车位置换业务: 车位置换测算表（非「运营」版本）
    has_voucher = 'calculation_tool' in result
    has_spot = 'discount_purchase' in result or 'spot_exchange_calc' in result
    if has_voucher and has_spot:
        result['business_type'] = 'both'
    elif has_voucher:
        result['business_type'] = 'parking_voucher'
    elif has_spot:
        result['business_type'] = 'spot_exchange'
    else:
        result['business_type'] = 'unknown'

    wb.close()
    return result


if __name__ == '__main__':
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else 'doc/2025-12-09【大良保利广场】停车券置换运营测算表更新版-承包方-快过期.xlsx'
    data = parse_audit_excel(path)
    print(json.dumps(data, ensure_ascii=False, indent=2))
